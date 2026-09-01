"""导入用户自己的框架。主 spec §7.3.5

安全团队手上的东西就是 Excel。所以吃 CSV / XLSX，不发明新格式。
"""
import pytest

from framework_reader.userframework.importer import ImportError_, parse_table, read_rows


def _rows(*rows):
    return [list(r) for r in rows]


def test_the_three_columns_are_id_label_parent():
    parsed = parse_table(_rows(["编号", "标题", "上级"], ["3.1", "账号管理", ""]))
    assert parsed == [("3.1", "账号管理", None, "")]


def test_english_headers_work_too():
    parsed = parse_table(_rows(["id", "label", "parent"], ["3.1", "Accounts", ""]))
    assert parsed == [("3.1", "Accounts", None, "")]


def test_a_parent_is_carried_through():
    parsed = parse_table(_rows(["编号", "标题", "上级"], ["3", "访问控制", ""], ["3.1", "账号", "3"]))
    assert parsed[1] == ("3.1", "账号", "3", "")


def test_the_parent_column_is_optional():
    assert parse_table(_rows(["编号", "标题"], ["3.1", "账号管理"])) == [("3.1", "账号管理", None, "")]


def test_a_missing_header_is_a_clear_error_not_a_crash():
    with pytest.raises(ImportError_, match="标题"):
        parse_table(_rows(["编号", "说明"], ["3.1", "账号管理"]))


def test_an_empty_file_is_rejected():
    with pytest.raises(ImportError_, match="empty"):
        parse_table([])


def test_rows_without_an_id_are_rejected_with_the_row_number():
    """静默跳过坏行，用户会以为全导进去了。"""
    with pytest.raises(ImportError_, match="Row 3"):
        parse_table(_rows(["编号", "标题"], ["3.1", "账号"], ["", "没有编号"]))


def test_duplicate_ids_are_rejected():
    with pytest.raises(ImportError_, match="duplicate"):
        parse_table(_rows(["编号", "标题"], ["3.1", "账号"], ["3.1", "又一个"]))


def test_a_parent_that_does_not_exist_is_rejected():
    with pytest.raises(ImportError_, match="parent"):
        parse_table(_rows(["编号", "标题", "上级"], ["3.1", "账号", "9.9"]))


def test_whitespace_is_stripped():
    assert parse_table(_rows(["编号", "标题"], ["  3.1 ", " 账号 "])) == [("3.1", "账号", None, "")]


def test_blank_lines_are_skipped():
    parsed = parse_table(_rows(["编号", "标题"], ["", ""], ["3.1", "账号"]))
    assert parsed == [("3.1", "账号", None, "")]


def test_csv_is_read(tmp_path):
    path = tmp_path / "f.csv"
    path.write_text("编号,标题\n3.1,账号管理\n", encoding="utf-8")
    assert parse_table(read_rows(path)) == [("3.1", "账号管理", None, "")]


def test_xlsx_is_read(tmp_path):
    from openpyxl import Workbook

    path = tmp_path / "f.xlsx"
    wb = Workbook()
    wb.active.append(["编号", "标题"])
    wb.active.append(["3.1", "账号管理"])
    wb.save(path)
    assert parse_table(read_rows(path)) == [("3.1", "账号管理", None, "")]


def test_an_unsupported_extension_says_what_is_supported(tmp_path):
    path = tmp_path / "f.docx"
    path.write_text("x", encoding="utf-8")
    with pytest.raises(ImportError_, match="csv"):
        read_rows(path)


# ---------- 正文列：用户自己的制度原文 ----------

def test_a_body_column_is_optional_and_carried_through():
    """用户自己公司的制度原文——他自己的文档、他自己的 key，可以拿去起草。"""
    parsed = parse_table(_rows(
        ["编号", "标题", "正文"],
        ["3.1", "账号管理", "各系统账号须由部门负责人审批后开立，离职当日回收。"],
    ))
    assert parsed[0][3].startswith("各系统账号")


def test_without_a_body_column_the_body_is_empty():
    assert parse_table(_rows(["编号", "标题"], ["3.1", "账号"]))[0][3] == ""


# ---------- 表头找不到时，别把人堵在死路上 ----------
#
# 「表头里找不到「编号」这一列」是对的，但它没说**它看见了什么**。
# 中文单位的表格最常见的情况是表头压在第二三行下面：第一行是个标题，
# 或者一片合并单元格。人看着自己的表明明有「编号」两个字，
# 只会觉得这工具坏了。

def test_the_error_shows_what_it_actually_found():
    with pytest.raises(ImportError_) as exc:
        parse_table(_rows(["公司信息安全控制清单", "", ""],
                          ["编号", "标题", "上级"],
                          ["3.1", "账号管理", ""]))
    assert "公司信息安全控制清单" in str(exc.value)


def test_a_header_further_down_is_pointed_out():
    """表头在第 2 行。说出来，人删掉上面那行就好了。"""
    with pytest.raises(ImportError_) as exc:
        parse_table(_rows(["公司信息安全控制清单", "", ""],
                          ["编号", "标题", "上级"],
                          ["3.1", "账号管理", ""]))
    assert "row 2" in str(exc.value)


def test_the_message_carries_no_markdown():
    """这些话会被 HTML 转义后渲到页面上。写 `**粗体**` 只会原样显示星号。"""
    with pytest.raises(ImportError_) as exc:
        parse_table(_rows(["标题行"], ["编号", "标题"], ["3.1", "账号管理"]))
    assert "**" not in str(exc.value)


def test_a_header_five_rows_down_is_found_too():
    with pytest.raises(ImportError_) as exc:
        parse_table(_rows(["标题行"], ["制表：安全部"], ["2026-03"], [""],
                          ["编号", "标题"], ["3.1", "账号管理"]))
    assert "row 5" in str(exc.value)


def test_no_header_anywhere_says_which_names_are_accepted():
    """真的没有表头时，得告诉人这一列可以叫什么。"""
    with pytest.raises(ImportError_) as exc:
        parse_table(_rows(["甲", "乙", "丙"], ["1", "2", "3"]))
    assert "控制编号" in str(exc.value) or "条号" in str(exc.value)


def test_a_header_in_the_first_row_still_just_works():
    """别为了报错好看，把正常路径搞出回归。"""
    parsed = parse_table(_rows(["编号", "标题"], ["3.1", "账号管理"]))
    assert parsed == [("3.1", "账号管理", None, "")]


# ---------- 一个工作簿里有好几张表 ----------
#
# 实测：aivtf-excel.xlsx 的**活动工作表是「Get started」说明页**，18 行全是
# 使用说明。真正的检查表在别的 sheet 里，而 `book.active` 只读一张——
# 模型看到一页说明文字，判定「这不是表」，它判断得对，错的是我们只给了一页。

def _workbook(sheets: dict) -> bytes:
    from io import BytesIO

    from openpyxl import Workbook

    book = Workbook()
    book.remove(book.active)
    for title, rows in sheets.items():
        sheet = book.create_sheet(title)
        for row in rows:
            sheet.append(row)
    buffer = BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def test_every_sheet_comes_back_with_its_name(tmp_path):
    from framework_reader.userframework.importer import read_sheets

    path = tmp_path / "book.xlsx"
    path.write_bytes(_workbook({
        "Get started": [["使用说明"], ["先看这里"]],
        "Checklist": [["编号", "标题"], ["3.1", "账号管理"]],
    }))
    sheets = read_sheets(path)
    assert [name for name, _ in sheets] == ["Get started", "Checklist"]
    assert sheets[1][1][0] == ["编号", "标题"]


def test_a_csv_is_one_nameless_sheet(tmp_path):
    from framework_reader.userframework.importer import read_sheets

    path = tmp_path / "x.csv"
    path.write_text("编号,标题\n3.1,账号管理\n", encoding="utf-8")
    sheets = read_sheets(path)
    assert len(sheets) == 1
    assert sheets[0][1][0] == ["编号", "标题"]


def test_the_first_sheet_that_parses_wins(tmp_path):
    """说明页在前、真表在后是最常见的排法。别停在第一张。"""
    from framework_reader.userframework.importer import (
        parse_any_sheet, read_sheets,
    )

    path = tmp_path / "book.xlsx"
    path.write_bytes(_workbook({
        "Get started": [["使用说明"], ["先看这里"]],
        "Checklist": [["编号", "标题"], ["3.1", "账号管理"]],
    }))
    name, controls = parse_any_sheet(read_sheets(path))
    assert name == "Checklist"
    assert controls == [("3.1", "账号管理", None, "")]


def test_no_sheet_parses_means_no_answer(tmp_path):
    from framework_reader.userframework.importer import (
        parse_any_sheet, read_sheets,
    )

    path = tmp_path / "book.xlsx"
    path.write_bytes(_workbook({
        "Get started": [["使用说明"]],
        "Notes": [["随手记"]],
    }))
    assert parse_any_sheet(read_sheets(path)) == (None, None)


def test_read_rows_still_works_for_the_single_sheet_callers(tmp_path):
    """CLI 那边还在用它，别把它拆没了。"""
    from framework_reader.userframework.importer import read_rows

    path = tmp_path / "x.csv"
    path.write_text("编号,标题\n3.1,账号管理\n", encoding="utf-8")
    assert read_rows(path)[0] == ["编号", "标题"]


# ---------- 删框架要删干净 ----------
#
# `remove()` 原来只删条款和解读，自评、签字、答题历史留在库里。而它自己的
# 注释说得很清楚：「留下够不着的孤儿行，下次导入同名框架还会串味。」
# 同一个理由对自评同样成立——删掉 ACME-1 再导一个同名的，
# 旧自评会自己长回来，而没人会想到去怀疑它。

def _framework_with_answers(tmp_path):
    from framework_reader.assess.store import AssessStore
    from framework_reader.userframework.store import UserFrameworkStore

    path = tmp_path / "user.sqlite"
    store = UserFrameworkStore(path)
    store.add_framework(framework_id="ACME-1", name="ACME",
                        controls=[("3.1", "账号管理", None, "正文")])
    AssessStore(path).record("ACME-1:3.1", level=2, note="有清单了")
    return path, store


def test_removing_a_framework_takes_its_assessments_with_it(tmp_path):
    from framework_reader.assess.store import AssessStore

    path, store = _framework_with_answers(tmp_path)
    store.remove("ACME-1")
    assert [a for a in AssessStore(path).all()
            if a.control_id.startswith("ACME-1:")] == []


def test_re_importing_the_same_id_does_not_inherit_the_old_answers(tmp_path):
    """这是「串味」的真样子：编号一样，答案是上一份文档的。"""
    from framework_reader.assess.store import AssessStore
    from framework_reader.userframework.store import UserFrameworkStore

    path, store = _framework_with_answers(tmp_path)
    store.remove("ACME-1")
    UserFrameworkStore(path).add_framework(
        framework_id="ACME-1", name="ACME 改", controls=[("3.1", "账号管理", None, "新正文")])
    assert [a for a in AssessStore(path).all()
            if a.control_id == "ACME-1:3.1"] == []


def test_removing_one_framework_leaves_the_others_alone(tmp_path):
    from framework_reader.assess.store import AssessStore
    from framework_reader.userframework.store import UserFrameworkStore

    path, store = _framework_with_answers(tmp_path)
    UserFrameworkStore(path).add_framework(
        framework_id="ACME-2", name="乙", controls=[("1.1", "别的", None, "")])
    AssessStore(path).record("ACME-2:1.1", level=1, note="别动我")
    store.remove("ACME-1")
    kept = [a for a in AssessStore(path).all() if a.control_id == "ACME-2:1.1"]
    assert len(kept) == 1


def test_the_counts_say_what_will_be_destroyed(tmp_path):
    """删之前要能告诉人「你会丢掉 N 条自评」——不然那是无声的破坏。"""
    from framework_reader.userframework.store import UserFrameworkStore

    path, store = _framework_with_answers(tmp_path)
    what = UserFrameworkStore(path).what_removing_costs("ACME-1")
    assert what["controls"] == 1
    assert what["assessments"] == 1


def test_the_counts_of_a_framework_that_is_not_there(tmp_path):
    from framework_reader.userframework.store import UserFrameworkStore

    what = UserFrameworkStore(tmp_path / "user.sqlite").what_removing_costs("X")
    assert what["controls"] == 0
