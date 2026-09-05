"""NIST CPRT cross-mapping import. spec §4.2①

Source: csf-2.0-to-sp800-53r5-mappings.xlsx (NIST OLIR #186, public domain)
Column and sheet names follow the real structure observed in Task 6 - see tests/fixtures/README.md.
"""
from pathlib import Path

from openpyxl import load_workbook

from framework_reader.schema.mapping import (
    Mapping,
    Provenance,
    ProvenanceLevel,
    Relation,
)

SOURCE_ID = "NIST-OLIR-csf-2.0-to-sp800-53r5"
SOURCE_VERSION = "2024-02"

# These two key names must match the real headers recorded in Task 6 verbatim.
COL_CSF = "Focal Document\nElement"
COL_53 = "Reference Document\nElement"


def parse_cprt_mappings(path: Path, sheet: str, header_row: int) -> list[Mapping]:
    wb = load_workbook(Path(path), read_only=True, data_only=True)
    ws = wb[sheet]

    rows = ws.iter_rows(min_row=header_row, values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)]
    idx_csf = header.index(COL_CSF)
    idx_53 = header.index(COL_53)

    edges: list[Mapping] = []
    for row in rows:
        csf = (row[idx_csf] or "").strip() if idx_csf < len(row) else ""
        ctl = (row[idx_53] or "").strip() if idx_53 < len(row) else ""
        if not csf or not ctl:
            continue
        from_id = f"NIST-CSF-2.0:{csf}"
        to_id = f"NIST-800-53-R5:{ctl.upper()}"
        if from_id == to_id:
            continue
        edges.append(
            Mapping(
                from_id=from_id,
                to_id=to_id,
                relation=Relation.RELATED,
                provenance=Provenance(
                    level=ProvenanceLevel.L1_OFFICIAL,
                    source=SOURCE_ID,
                    source_version=SOURCE_VERSION,
                ),
                note="",
            )
        )
    return edges
