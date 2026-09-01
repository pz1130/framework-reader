"""ISO 27002:2022 骨架导入。spec §4.1 Tier C、§4.2①

只导入编号与自写 label。原文一个字都不进本仓库。
"""
import csv
import re
from pathlib import Path

from openpyxl import load_workbook

from framework_reader.schema.entities import Framework, FrameworkControl, LicenseTier
from framework_reader.schema.mapping import (
    Mapping,
    Provenance,
    ProvenanceLevel,
    Relation,
)

FRAMEWORK_ID = "ISO-27002-2022"
ISO_MAP_SOURCE = "NIST-SP800-53r5-to-iso-27001"
ISO_MAP_VERSION = "rev5-upd1"

# Observed OLIR #155 headers (tests/fixtures/README.md §5).
COL_53 = "Focal Document\nElement"
COL_ISO = "Reference Document Element"

# ISO/IEC 27001:2022 management-system clauses. Bare numbers matching these
# are not Annex A; A.-prefixed refs stay on the 27002 skeleton.
_ISO27001_CLAUSES = frozenset(
    {
        "4.1",
        "4.2",
        "4.3",
        "4.4",
        "5.1",
        "5.2",
        "5.3",
        "6.1",
        "6.1.1",
        "6.1.2",
        "6.1.3",
        "6.2",
        "6.3",
        "7.1",
        "7.2",
        "7.3",
        "7.4",
        "7.5",
        "7.5.1",
        "7.5.2",
        "7.5.3",
        "8.1",
        "8.2",
        "8.3",
        "9.1",
        "9.2",
        "9.2.1",
        "9.2.2",
        "9.3",
        "9.3.1",
        "9.3.2",
        "9.3.3",
        "10.1",
        "10.2",
    }
)

# ISO/IEC 27002:2022 Annex A leaf identifiers (factual numbers).
_ANNEX_A_LEAVES = frozenset(
    [f"A.5.{i}" for i in range(1, 38)]
    + [f"A.6.{i}" for i in range(1, 9)]
    + [f"A.7.{i}" for i in range(1, 15)]
    + [f"A.8.{i}" for i in range(1, 35)]
)

_ISO_REF_RE = re.compile(
    r"(?:(?P<annex>Annex\s+A\.?\s*)|(?P<a>A\.))?(?P<num>\d+\.\d+(?:\.\d+)?)",
    re.IGNORECASE,
)


def parse_iso_skeleton(path: Path) -> tuple[Framework, list[FrameworkControl]]:
    framework = Framework(
        id=FRAMEWORK_ID,
        name="ISO/IEC 27002:2022",
        version="2022",
        tier=LicenseTier.C_PURCHASE,
        source_url="https://www.iso.org/standard/75652.html",
        license_note="Purchase required; the original text must not be redistributed. The product stores only ids and self-written labels",
    )
    controls: list[FrameworkControl] = []
    with Path(path).open(encoding="utf-8", newline="") as fh:
        for row in csv.DictReader(fh):
            parent = row["parent_id"].strip()
            controls.append(
                FrameworkControl(
                    id=f"{FRAMEWORK_ID}:{row['control_id'].strip()}",
                    framework_id=FRAMEWORK_ID,
                    parent_id=f"{FRAMEWORK_ID}:{parent}" if parent else None,
                    label=row["label_zh"].strip(),
                    label_is_original=False,  # Tier C：必须自写
                    framework_tier=LicenseTier.C_PURCHASE,
                )
            )
    return framework, controls


def parse_800_53_to_iso(path: Path) -> list[Mapping]:
    """从 NIST 署名的 800-53 ↔ ISO 27001 对照表解析 L1 边。

    源文件为 OLIR #155 xlsx（原随附 docx 已 404）。只读 800-53 与 ISO
    编号列，不读描述列。Annex A 端点落到骨架命名空间 ISO-27002-2022:。
    """
    edges: list[Mapping] = []
    for ctl_53, iso_ref in _read_xlsx_id_pairs(path):
        to_id = _normalize_iso_endpoint(iso_ref)
        if to_id is None:
            continue
        edges.append(
            Mapping(
                from_id=f"NIST-800-53-R5:{ctl_53.upper()}",
                to_id=to_id,
                relation=Relation.RELATED,
                provenance=Provenance(
                    level=ProvenanceLevel.L1_OFFICIAL,
                    source=ISO_MAP_SOURCE,
                    source_version=ISO_MAP_VERSION,
                ),
                note="",
            )
        )
    return edges


def _cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _read_xlsx_id_pairs(path: Path) -> list[tuple[str, str]]:
    """读取各家族 sheet 的 800-53 / ISO 编号列；跳过空行与无对应表头的 sheet。"""
    wb = load_workbook(Path(path), read_only=True, data_only=True)
    out: list[tuple[str, str]] = []
    for name in wb.sheetnames:
        if name == "Definitions":
            continue
        ws = wb[name]
        rows = ws.iter_rows(min_row=1, values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            continue
        header_cells = [_cell(c) for c in header]
        try:
            idx_53 = header_cells.index(COL_53)
            idx_iso = header_cells.index(COL_ISO)
        except ValueError:
            continue
        for row in rows:
            ctl = _cell(row[idx_53]) if idx_53 < len(row) else ""
            iso = _cell(row[idx_iso]) if idx_iso < len(row) else ""
            if not ctl or not iso:
                continue
            out.append((ctl, iso))
    return out


def _normalize_iso_endpoint(raw: str) -> str | None:
    """把 A.8.16 / 8.16 / Annex A.8.16 归一到骨架 ID；跳过无骨架的 27001 条款。"""
    text = raw.strip()
    if not text:
        return None
    match = _ISO_REF_RE.fullmatch(text)
    if not match:
        return None
    num = match.group("num")
    had_annex_prefix = bool(match.group("annex") or match.group("a"))
    if not had_annex_prefix and num in _ISO27001_CLAUSES:
        return None
    control_id = f"A.{num}"
    if control_id not in _ANNEX_A_LEAVES:
        return None
    return f"{FRAMEWORK_ID}:{control_id}"
