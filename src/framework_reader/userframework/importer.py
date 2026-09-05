"""Read CSV / XLSX, emit (number, title, parent) triples. Main spec §7.3.5

What security teams actually have in hand is Excel, so no new format is invented.
**A bad row is always an error naming its line number, never silently skipped** -
silent skipping ends with the user believing everything was imported.
"""
import csv
from pathlib import Path

_ID_HEADERS = {"编号", "控制编号", "条号", "id", "control_id", "ref", "ref_id"}
_LABEL_HEADERS = {"标题", "名称", "控制", "条款", "label", "name", "title"}
_PARENT_HEADERS = {"上级", "父级", "上级编号", "parent", "parent_id"}
# The user's own company policy text. His document, his machine, his key - it may
# be sent out for drafting. That is a completely different thing from the
# copyrighted standard text of Tier C/D, which may never leave the network
# (main spec §9).
_BODY_HEADERS = {"正文", "描述", "要求", "内容", "条款正文", "body", "text", "description"}
MAX_SHEETS = 100
MAX_ROWS = 100_000
MAX_CELLS = 2_000_000


class ImportError_(Exception):
    """Import failed. The message must let the user fix the sheet himself, so it
    always carries a line number or a column name."""


def _bounded_rows(rows) -> list[list[str]]:
    out = []
    cells = 0
    for number, row in enumerate(rows, start=1):
        if number > MAX_ROWS:
            raise ImportError_(f"The file has more than {MAX_ROWS:,} rows")
        values = ["" if cell is None else str(cell) for cell in row]
        cells += len(values)
        if cells > MAX_CELLS:
            raise ImportError_(f"The file has more than {MAX_CELLS:,} cells")
        out.append(values)
    return out


def read_sheets(path: Path) -> list[tuple[str, list[list[str]]]]:
    """**Every sheet** in a workbook, with its name.

    `book.active` reads only one sheet, and "instructions first, real table
    after" is the most common layout - one checklist workbook observed in
    practice had "Get started" as its active sheet, 18 rows of usage
    instructions, with the real checklist in another sheet. Reading only one
    sheet wastes the whole import.
    """
    suffix = Path(path).suffix.lower()
    if suffix == ".csv":
        with Path(path).open(encoding="utf-8-sig", newline="") as handle:
            return [("", _bounded_rows(csv.reader(handle)))]
    if suffix in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook

        book = load_workbook(path, read_only=True, data_only=True)
        try:
            if len(book.worksheets) > MAX_SHEETS:
                raise ImportError_(f"The workbook has more than {MAX_SHEETS} sheets")
            return [
                (sheet.title, _bounded_rows(sheet.iter_rows(values_only=True)))
                for sheet in book.worksheets
            ]
        finally:
            book.close()
    raise ImportError_(f"Unknown file type {suffix} - supported: .csv and .xlsx")


def read_rows(path: Path) -> list[list[str]]:
    """The first sheet. The CLI and existing callers still use it."""
    sheets = read_sheets(path)
    return sheets[0][1] if sheets else []


def parse_any_sheet(
    sheets: list[tuple[str, list[list[str]]]]
) -> tuple[str | None, list[tuple[str, str, str | None, str]] | None]:
    """Try them one by one; the first sheet that parses wins. If none parse,
    return (None, None) - that is when the model should take a look at the whole
    workbook (see `_shape_table` in `web/app.py`).
    """
    for name, rows in sheets:
        try:
            return name, parse_table(rows)
        except ImportError_:
            continue
    return None, None


def _column(header: list[str], names: set[str], what: str, required: bool = True,
            rows: list[list[str]] | None = None) -> int | None:
    for index, cell in enumerate(header):
        if str(cell).strip().lower() in names:
            return index
    if required:
        raise ImportError_(_missing_header_message(what, names, header, rows or []))
    return None


# How many rows down to search. Tables from Chinese organizations often have a
# title row or a stretch of merged cells sitting on top of the header; any deeper
# than that and it is no longer "header pushed down" but some other problem.
_HEADER_SEARCH_ROWS = 10


def _missing_header_message(what: str, names: set[str],
                            header: list[str], rows: list[list[str]]) -> str:
    """**Say what it actually saw.** If the message only says "the number column
    was not found" while the person is looking straight at the two characters for
    "number" in his own sheet, he will conclude the tool is broken.
    """
    seen = "、".join(str(c).strip() for c in header if str(c).strip()) or "(empty)"
    for number, row in enumerate(rows[1:_HEADER_SEARCH_ROWS], start=2):
        if any(str(cell).strip().lower() in names for cell in row):
            return (
                f'Column "{what}" not found in the header - the first row is: {seen[:60]}.'
                # This sentence is HTML-escaped when rendered onto the page; markdown would just show asterisks literally.
                f"The real header looks like it is on row {number}, "
                "delete the rows above it and upload again."
            )
    accepted = ", ".join(sorted(n for n in names if not n.isascii()))
    return (
        f'Column "{what}" not found in the header. The first row is: {seen[:60]}.'
        f'This column can be named: {accepted}. The first row also needs "Title" ("Parent" and "Body" optional).'
    )
    return None


def parse_table(rows: list[list[str]]) -> list[tuple[str, str, str | None, str]]:
    if not rows:
        raise ImportError_("The file is empty")
    header = rows[0]
    id_col = _column(header, _ID_HEADERS, "编号", rows=rows)
    label_col = _column(header, _LABEL_HEADERS, "标题", rows=rows)
    parent_col = _column(header, _PARENT_HEADERS, "上级", required=False)
    body_col = _column(header, _BODY_HEADERS, "正文", required=False)

    out: list[tuple[str, str, str | None, str]] = []
    seen: set[str] = set()
    for number, row in enumerate(rows[1:], start=2):
        def cell(index: int | None) -> str:
            if index is None or index >= len(row):
                return ""
            return str(row[index]).strip()

        local, label, parent = cell(id_col), cell(label_col), cell(parent_col)
        if not local and not label:
            continue          # entirely blank row, skip
        if not local:
            raise ImportError_(f"Row {number} has no number (title: '{label[:20]}')")
        if not label:
            raise ImportError_(f"Row {number} has no title (number: {local})")
        if local in seen:
            raise ImportError_(f"Row {number} has duplicate number {local}")
        seen.add(local)
        out.append((local, label, parent or None, cell(body_col)))

    if not out:
        raise ImportError_("No data rows below the header")

    known = {local for local, _label, _parent, _body in out}
    for local, _label, parent, _body in out:
        if parent and parent not in known:
            raise ImportError_(f"{local} has parent {parent} which is not in the table")
    return out
